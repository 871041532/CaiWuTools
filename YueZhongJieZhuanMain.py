
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *
import sys
import getpass
from collections import OrderedDict
from pyexcel_xlsx import get_data, save_data
from pyexcel_xls import get_data
import csv
import re
import copy
import datetime
from copy import deepcopy
from Globals import Globals
import datetime
import openpyxl
import shutil
import os

class ShowWindow(QWidget):
    LOG_WORLD = 1
    LOG_ERROR = 2

    def closeEvent(self, QCloseEvent):
        if hasattr(Globals, "MainWin"):
            Globals.current_win = Globals.MainWin()

    """docstring for ClassName"""

    def __init__(self):
        super(ShowWindow, self).__init__()
        self.initUI()

    def initUI(self):
        mainLayout = QVBoxLayout()

        self.button1 = QPushButton("选择海鼎导出结转表")
        self.button1.clicked.connect(self.click_select_my)
        self.button1.setMinimumHeight(70)

        self.button2 = QPushButton("选择需要填充的7个结转模板")
        self.button2.clicked.connect(self.click_select_output)
        self.button2.setMinimumHeight(70)

        cur = datetime.datetime.now()
        self.label = QLineEdit(str(cur.month) + "月")
        self.label.setPlaceholderText("输入结转模板sheet名字 如: 1月")
        self.label.setMinimumHeight(50)

        self.button3 = QPushButton("开始填充结转模板")
        self.button3.clicked.connect(self.click_judge_btn)
        self.button3.setMinimumHeight(70)

        self.button4 = QPushButton("生成主合同结转凭证")
        self.button4.clicked.connect(self.click_deal_pingzheng)
        self.button4.setMinimumHeight(70)

        self.button5 = QPushButton("选择莘宝结转基础信息表")
        self.button5.setMinimumHeight(70)
        self.button5.clicked.connect(self.click_select_refer)

        self.text_browser = QTextBrowser()
        mainLayout.addWidget(self.button1)
        mainLayout.addWidget(self.button2)
        mainLayout.addWidget(self.label)
        mainLayout.addWidget(self.button3)
        mainLayout.addWidget(self.button5)
        mainLayout.addWidget(self.button4)
        mainLayout.addWidget(self.text_browser)

        # data
        self.my_filename = None # Globals.desktop_path + "原始表.xls"
        self.out_put_filenames = None
        self.refer_filename = None # Globals.desktop_path + "莘宝结转基础信息.xlsx"
        self.refer_data = None
        self.my_data = None
        self.out_put_datas = None

        self.setLayout(mainLayout)
        self.setWindowTitle("月中结转主合同")
        self.setGeometry(550, 150, 900, 800)
        self.show()

        if not Globals.user:
            self.close()

    # 选择目标表
    def click_select_output(self):
        info = QFileDialog.getOpenFileNames(self, '选择需要填充的7个结转模板')
        if info and info[0]:
            self.out_put_filenames = info[0]
            self.log("选择了结转模板：" + str(self.out_put_filenames))

    # 选择基础信息表
    def click_select_refer(self):
        info = QFileDialog.getOpenFileName(self, '选择莘宝结转基础信息表')
        if info and info[0]:
            self.refer_filename = info[0]
            self.log("选择莘宝结转基础信息表：" + self.refer_filename)

    # 选择源
    def click_select_my(self):
        info = QFileDialog.getOpenFileName(self, '选择海鼎导出结转表')
        if info and info[0]:
            self.my_filename = info[0]
            self.log("选择海鼎导出结转表：" + self.my_filename)

    def clear_log(self):
        self.text_browser.clear()

    def click_judge_btn(self):
        self.clear_log()
        if not (self.my_filename):
            self.log("请先选择海鼎导出结转表~(^v^)~", ShowWindow.LOG_ERROR)
        elif not self.label.text():
            self.log("请先输入月份！", ShowWindow.LOG_ERROR)
        elif not self.out_put_filenames:
            self.log("请先选择一点几输出文件~(^v^)~", ShowWindow.LOG_ERROR)
        else:
            self.load_my_data()
            self.load_muban_data()
            self.deal_muban_data()
            self.log("处理完毕。")

    def click_deal_pingzheng(self):
        self.clear_log()
        if not (self.my_filename):
            self.log("请先选择海鼎导出结转表~(^v^)~", ShowWindow.LOG_ERROR)
        elif not self.refer_filename:
            self.log("选择莘宝结转基础信息表~(^v^)~", ShowWindow.LOG_ERROR)
        else:
            self.load_my_data()
            self.load_refer_data()
            self.deal_pingzheng_data()
            self.log("处理完毕。")

    def load_my_data(self):
        self.log(self.my_filename.split('/')[-1] + " 导入成功。")
        self.my_data = get_data(self.my_filename)
        self.my_data = list(self.my_data.values())[0]
        self.my_data = self.my_data[1:]
        row_0 = self.my_data[0]
        row_1 = self.my_data[1]
        for i in range(len(row_1)):
            if row_0[i] == "":
                row_0[i] = row_1[i]
        self.my_data = self.my_data[1:]
        self.my_data[0] = row_0

    def load_refer_data(self):
        self.log(self.refer_filename.split('/')[-1] + " 导入成功。")
        self.refer_data = get_data(self.refer_filename)["EAS基础信息"][1:]
        dict = {}
        for row in self.refer_data:
            dict[row[1]] = row
        self.refer_data = dict

    def load_muban_data(self):
        self.out_put_datas = {}
        for full_path in self.out_put_filenames:
            file_name = full_path.split('/')[-1]
            self.out_put_datas[file_name] = (full_path, openpyxl.load_workbook(full_path))
            self.log(file_name + " 导入成功。")

    # 处理凭证数据
    def deal_pingzheng_data(self):
        # 先检测一遍基础信息是否齐全
        gongsimingcheng_idx = self.my_data[0].index("公司名称")
        is_ok = True
        for i in range(1, len(self.my_data)):
            if len(self.my_data[i]) > gongsimingcheng_idx and self.my_data[i][gongsimingcheng_idx] and self.my_data[i][0]:
                if self.my_data[i][gongsimingcheng_idx] not in self.refer_data:
                    self.log("公司: " + self.my_data[i][gongsimingcheng_idx] + " 不在基础信息表中", self.LOG_ERROR)
                    is_ok = False
        if not is_ok:
            self.log("一定要注意删掉禁用的自定义核算项目！",self.LOG_ERROR)
            return
        else:
            self.log("公司名称检测ok!")

        # 基础信息
        out_put_file_dir = Globals.desktop_path + "主合同结转引入凭证/"
        if not os.path.exists(out_put_file_dir):
            os.mkdir(out_put_file_dir)
        cur = datetime.datetime.now()
        kemu_date_str = str(cur.year) + "%02d"%cur.month
        kemu_date_str_last_month = Globals.get_time_text_year_lastmonth()
        output_filenames = (
            # key，税率，非主力店科目，主力店科目
            ("仓库租赁费", 1.1, "6001.05", "6001.05", "结转" + kemu_date_str + "仓库租赁费收入 %s-%s-%s"),
            ("浮动提成租金", 1.1, "6001.02.02", "6001.02.01", "结转" + kemu_date_str_last_month + "浮动提成租金 %s-%s-%s"),
            ("固定租金", 1.1, "6001.01.02", "6001.01.01", "结转" + kemu_date_str + "固定租金 %s-%s-%s"),
            ("广告位租赁费", 1.1, "6001.08", "6001.08", "结转" + kemu_date_str + "广告位租赁费收入 %s-%s-%s"),
            ("推广费(销售提成）", 1.06, "6001.12", "6001.12", "结转" + kemu_date_str_last_month + "推广费(销售提成） %s-%s-%s"),
            ("推广费（固定）", 1.06, "6001.12", "6001.12", "结转" + kemu_date_str + "推广费（固定） %s-%s-%s"),
            ("物业管理费", 1.06, "6001.04.02", "6001.04.01", "结转" + kemu_date_str + "物业管理费收入 %s-%s-%s"),
        )
        for file_info in output_filenames:
            # 主key
            file_name = file_info[0]
            # 税率
            shuilv = file_info[1]
            # 非主力科目
            not_zhuli_kemu = file_info[2]
            # 主力科目
            zhuli_kemu = file_info[3]
            # 摘要str
            zhaiyao_str = file_info[4]
            # 获取demo
            excel_data = deepcopy(Globals.get_origin_excel_data())
            target_file_name = out_put_file_dir + "结转主合同" + file_name + ".xlsx"
            title_key = file_name
            title_idx = self.my_data[0].index(title_key)
            origin_data = []
            for i in range(1, len(self.my_data)):
                if len(self.my_data[i]) > title_idx and self.my_data[i][title_idx] and self.my_data[i][0]:
                    origin_data.append(self.my_data[i])

            dianpuzhaopai_idx = self.my_data[0].index("店铺招牌")
            puweihao_idx = self.my_data[0].index("铺位号")
            def key_func(elem):
                return elem[puweihao_idx]
            origin_data.sort(key=key_func)
            target_rows = []
            data_str = str(cur.year) + "/" + str(cur.month) + "/" +str(cur.day)
            # 处理数据
            for i in range(len(origin_data)):
                one_data = origin_data[i]
                row = deepcopy(Globals.pingzheng_demo)
                row2 = deepcopy(Globals.pingzheng_demo)
                target_rows.append(row)
                target_rows.append(row2)
                row[Globals.get_pingzheng_idx("记账日期")] = data_str
                row2[Globals.get_pingzheng_idx("记账日期")] = data_str
                row[Globals.get_pingzheng_idx("业务日期")] = data_str
                row2[Globals.get_pingzheng_idx("业务日期")] = data_str
                row[Globals.get_pingzheng_idx("辅助账业务日期")] = data_str
                row2[Globals.get_pingzheng_idx("辅助账业务日期")] = data_str
                row[Globals.get_pingzheng_idx("会计期间")] = str(cur.month)
                row2[Globals.get_pingzheng_idx("会计期间")] = str(cur.month)
                row[Globals.get_pingzheng_idx("凭证类型")] = "转"
                row2[Globals.get_pingzheng_idx("凭证类型")] = "转"
                row[Globals.get_pingzheng_idx("凭证号")] = "20180600269"
                row2[Globals.get_pingzheng_idx("凭证号")] = "20180600269"
                row[Globals.get_pingzheng_idx("分录号")] = i * 2 + 1
                row2[Globals.get_pingzheng_idx("分录号")] = i * 2 + 2
                shop_name = one_data[gongsimingcheng_idx]
                zhiapai_name = one_data[dianpuzhaopai_idx]
                puweihao_name = one_data[puweihao_idx]
                zhaiyao = zhaiyao_str%(shop_name, zhiapai_name, puweihao_name)
                row[Globals.get_pingzheng_idx("摘要")] = zhaiyao
                row2[Globals.get_pingzheng_idx("摘要")] = zhaiyao
                row[Globals.get_pingzheng_idx("科目")] = "2203.01.01"
                if self.is_zhuli(shop_name):
                    row2[Globals.get_pingzheng_idx("科目")] = zhuli_kemu
                else:
                    row2[Globals.get_pingzheng_idx("科目")] = not_zhuli_kemu
                row[Globals.get_pingzheng_idx("方向")] = 1
                row2[Globals.get_pingzheng_idx("方向")] = 0
                jine_num = float("%.2f" % (one_data[title_idx] / shuilv))
                row[Globals.get_pingzheng_idx("原币金额")] = jine_num
                row2[Globals.get_pingzheng_idx("原币金额")] = jine_num
                row[Globals.get_pingzheng_idx("借方金额")] = jine_num
                row2[Globals.get_pingzheng_idx("贷方金额")] = jine_num
                row[Globals.get_pingzheng_idx("现金流量标记")] = 2
                row2[Globals.get_pingzheng_idx("现金流量标记")] = 2
                row[Globals.get_pingzheng_idx("辅助账摘要")] = zhaiyao_str
                row[Globals.get_pingzheng_idx("核算项目1")] = "长益租户"
                row[Globals.get_pingzheng_idx("名称1")] = shop_name
                row[Globals.get_pingzheng_idx("编码1")] = self.get_bianma(shop_name)



            for row in target_rows:
                excel_data["凭证"].append(row)
            save_data(target_file_name, excel_data)
            Globals.eval_date_format(target_file_name, ["B", "C", "BN"])

    # 判断是不是主力店
    def is_zhuli(self, shop_name):
        return self.refer_data[shop_name][4] == "主力商户"

    # 获得编码
    def get_bianma(self, shop_name):
        return self.refer_data[shop_name][2]

    # 处理模板数据
    def deal_muban_data(self):
        # 处理 1.1
        for k, v in self.out_put_datas.items():
            title_key = ""
            if "1.1" == k[0:3]:
                # 获取原始数据
                title_key = "仓库租赁费"
            elif "1.2" == k[0:3]:
                title_key = "浮动提成租金"
            elif "1.3" == k[0:3]:
                title_key = "固定租金"
            elif "1.4" == k[0:3]:
                title_key = "广告位租赁费"
            elif "1.5" == k[0:3]:
                title_key = "推广费(销售提成）"
            elif "1.6" == k[0:3]:
                title_key = "推广费（固定）"
            elif "1.7" == k[0:3]:
                title_key = "物业管理费"
            if title_key:
                title_idx = self.my_data[0].index(title_key)
                origin_data = []
                for i in range(1, len(self.my_data)):
                    if len(self.my_data[i]) > title_idx and self.my_data[i][title_idx] and self.my_data[i][0]:
                        origin_data.append(self.my_data[i])
                gongsimingcheng_idx = self.my_data[0].index("公司名称")
                dianpuzhaopai_idx = self.my_data[0].index("店铺招牌")
                puweihao_idx = self.my_data[0].index("铺位号")
                def key_func(elem):
                    return elem[puweihao_idx]
                origin_data.sort(key = key_func)
                # 获取目标数据
                excel = v[1]
                sheet = excel[self.label.text()]
                # 写入
                for i in range(len(origin_data)):
                    save_row_idx = i + 3
                    # 租户名称
                    sheet.cell(row = save_row_idx, column = 1).value = origin_data[i][gongsimingcheng_idx]
                    # 店铺招牌
                    sheet.cell(row = save_row_idx, column = 2).value = origin_data[i][dianpuzhaopai_idx]
                    # 铺位号
                    sheet.cell(row = save_row_idx, column = 3).value = origin_data[i][puweihao_idx]
                    # 金额
                    sheet.cell(row=save_row_idx, column = 4).value = origin_data[i][title_idx]
                v[1].save(v[0])

    def log(self, strs, enum=1):
        if enum == self.LOG_ERROR:
            pre_str = """<font face="微软雅黑" size="6" color="red">%s</font>"""
            self.text_browser.append(pre_str % str(strs))
        else:
            pre_str = """<font face="微软雅黑" size="4" color="black">%s</font>"""
            self.text_browser.append(pre_str % str(strs))


def into():
    return ShowWindow()


if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = ShowWindow()
    sys.exit(app.exec_())
